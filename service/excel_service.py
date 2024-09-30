from pdb import run
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import pandas as pd
from global_constant import GLOBAL_CONSTANT
from model.report_model import ReportModel

# ... (Assume you have 'list_report_model' containing your data)

def create_excel_report(list_report_model:list[ReportModel]):

    # Extract data for the first sheet (RawReportModel)
    for report_model in list_report_model:
        # Extract data for the raw data sheet (now using raw_report_model_ip)
        raw_data = []
        for raw_model_ip in report_model.raw_report_model_ip:
            # Filter out non-SMB data if the report is an SMB report
            raw_data.append({
                'ID': raw_model_ip.id,
                'log_date': raw_model_ip.log_date,
                'srcip': raw_model_ip.scr_ip,
                'dstport': raw_model_ip.dst_port,
                'Application': raw_model_ip.application,
                'total': raw_model_ip.total
            })

        # Create DataFrame for the raw data sheet
        df_raw = pd.DataFrame(raw_data)

        # Extract data for the pivot sheet
        pivot_data = []
        if report_model.report_model_list_ip_pivot:
            running_number = 1
            for pivot_model in report_model.report_model_list_ip_pivot.report_model_ip_pivot:
                list_date = report_model.report_model_list_ip_pivot.list_date
                row = {'No': running_number,'ip': pivot_model.ip}
                for date in list_date:
                    date_total = pivot_model.date_total.get(date, 0) 
                    row[date] = date_total

                row['Sum_total'] = pivot_model.total  # OR calculate sum here if `pivot_model.total` is not available
                pivot_data.append(row)  # Append row after calculating Sum_total
                running_number += 1

        df_pivot = pd.DataFrame(pivot_data)
        df_final = None

        if not df_pivot.empty:
            # df_pivot['Sum_total'] = df_pivot.iloc[:, 2:-1].sum(axis=1)  # No need to recalculate if already done above
            df_pivot = df_pivot.sort_values(by='Sum_total', ascending=False)
        
        df_pivot[" "] = ""
        if report_model.report_model_list_ip_pivot:
                # Create a list to store data for the new DataFrame
                top_20_data = []
                
                running_number = 1
                for pivot_model in report_model.report_model_list_ip_pivot.top_20_ip_pivot:
                    list_date = report_model.report_model_list_ip_pivot.list_date
                    row = {'No': running_number, 'IP': pivot_model.ip}
                    if 'Branch' in report_model.report_name:
                        row['Subnet IP'] = pivot_model.subnet_ip
                        row['Branch'] = pivot_model.subnet_branch
                    running_number += 1
                    top_20_data.append(row)

                # Create a DataFrame for top_20_ip_pivot data
                df_top_20 = pd.DataFrame(top_20_data)
                df_final = pd.concat([df_pivot, pd.DataFrame({' ': ''}, index=df_pivot.index), df_top_20], axis=1)

        # Create Excel writer with dynamic filename per report
        filename = f"({GLOBAL_CONSTANT.START_TIME.strftime('%Y-%m-%d')}_to_{GLOBAL_CONSTANT.END_TIME.strftime('%Y-%m-%d')})_{report_model.report_title}.xlsx"

        with pd.ExcelWriter(filename) as writer:
            # Write data to the raw data sheet
            df_raw.to_excel(writer, sheet_name='Raw Data', index=False)

            # Write data to the pivot sheet
            if df_final is not None:
                df_final.to_excel(writer, sheet_name='Pivot Data', index=False)
            else:
                df_pivot.to_excel(writer, sheet_name='Pivot Data', index=False)

            # Auto-adjust column widths for both sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try: 
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2) 
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        if 'Branch' in report_model.report_name:
            workbook = load_workbook(filename)

            for sheet_name in ['Pivot Data']:
                sheet = workbook[sheet_name]

            sum_total_index = None
            for col_idx, col in enumerate(sheet.iter_cols(1, sheet.max_column)):  # Iterate through columns
                if col[0].value == 'Sum_total':
                    sum_total_index = col_idx
                    break
                # Set header fill color (assuming header is in the first row)
            blue_fill = PatternFill(start_color="FF809FFF", end_color="FF809FFF", fill_type="solid")
            if sum_total_index is not None:
                if 'Branch' in report_model.report_name:
                    for col_idx in range(sum_total_index + 4, sum_total_index + 8):  # 3 columns after 'Sum_total'
                        cell = sheet.cell(row=1, column=col_idx) 
                        cell.fill = blue_fill
                # else :
                #     for col_idx in range(sum_total_index + 4, sum_total_index + 5):  # 3 columns after 'Sum_total'
                #         cell = sheet.cell(row=1, column=col_idx) 
                #         cell.fill = blue_fill
                    # Remove gridlines
                    sheet.sheet_view.showGridLines = False
                workbook.save(filename)
        print(f"Excel report '{filename}' created successfully!")